"""
NFX Signal xlsx ingestion adapter.

Parses investor lists exported from https://signal.nfx.com — the NFX VC network
platform. Detected by the presence of columns: "Investor Full Name" and
"Signal Profile link".

Produces NfxInvestorRecord objects (via extract_records) for direct use by the
batch gate runner, and RawRecord objects (via extract_raw) for the standard
ingestion pipeline.
"""

from __future__ import annotations

import hashlib
import re
from pathlib import Path
from typing import Iterable, Iterator, List, Optional

import openpyxl

from agents.ingestion.base import (
    IngestionAdapter,
    RawRecord,
    Source,
    SourceManifest,
    hash_content,
    make_source_record_id,
    should_skip_raw_file,
)
from contra.gate.batch_models import NfxInvestorRecord


# ---------------------------------------------------------------------------
# Detection helpers
# ---------------------------------------------------------------------------

_NFX_REQUIRED_COLUMNS = {"investor full name", "signal profile link"}
_NFX_SIGNAL_DOMAIN = "signal.nfx.com"


def _is_nfx_xlsx(path: Path) -> bool:
    """Heuristic detection: check first row for NFX Signal column names."""
    if should_skip_raw_file(path):
        return False
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True, max_row=1))
            if not rows:
                continue
            header = {str(c).strip().lower() for c in rows[0] if c is not None}
            if _NFX_REQUIRED_COLUMNS.issubset(header):
                wb.close()
                return True
        wb.close()
    except Exception:
        pass
    return False


def _clean(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _clean_dollar(val) -> Optional[str]:
    """Normalise dollar strings — keep as-is (already '$50,000' style)."""
    cleaned = _clean(val)
    if not cleaned:
        return None
    # Strip any stray whitespace inside dollar amounts
    cleaned = re.sub(r"\s+", "", cleaned)
    return cleaned


# ---------------------------------------------------------------------------
# Adapter
# ---------------------------------------------------------------------------

class NfxXlsxAdapter(IngestionAdapter):
    """Ingests NFX Signal xlsx investor exports."""

    source_type = "signal-nfx"
    schema_version = "1.0"

    # Column name aliases (lower-cased)
    _COL_MAP = {
        "firm name": "firm_name",
        "investor full name": "investor_name",
        "signal profile link": "nfx_url",
        "sweet spot": "sweet_spot",
        "min": "check_min",
        "max": "check_max",
        "locations": "locations",
        "intro source": "intro_source",
        "intro strength": "intro_strength",
    }

    def discover(self, root: Path) -> Iterable[Source]:
        for path in sorted(root.glob("**/*.xlsx")):
            if not _is_nfx_xlsx(path):
                continue
            file_hash = hashlib.sha256(path.read_bytes()).hexdigest()
            yield Source(
                path=path,
                source_type=self.source_type,
                relative_path=str(path.relative_to(root)),
                file_hash=file_hash,
                metadata={"source_platform": "nfx_signal"},
            )

    def extract_records(self, source: Source) -> List[NfxInvestorRecord]:
        """Parse xlsx rows into NfxInvestorRecord objects for the batch runner."""
        records: List[NfxInvestorRecord] = []
        wb = openpyxl.load_workbook(source.path, read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    continue

                # Build column index map from header row
                header = all_rows[0]
                col_idx: dict[str, int] = {}
                for i, col in enumerate(header):
                    if col is None:
                        continue
                    key = str(col).strip().lower()
                    if key in self._COL_MAP:
                        col_idx[self._COL_MAP[key]] = i

                if "investor_name" not in col_idx:
                    continue

                for row in all_rows[1:]:
                    def get(field: str) -> Optional[str]:
                        idx = col_idx.get(field)
                        if idx is None or idx >= len(row):
                            return None
                        return _clean(row[idx])

                    name = get("investor_name")
                    if not name:
                        continue

                    records.append(
                        NfxInvestorRecord(
                            investor_name=name,
                            firm_name=get("firm_name"),
                            nfx_url=get("nfx_url"),
                            sweet_spot=_clean_dollar(get("sweet_spot")),
                            check_min=_clean_dollar(get("check_min")),
                            check_max=_clean_dollar(get("check_max")),
                            locations=get("locations"),
                            intro_source=get("intro_source"),
                            intro_strength=get("intro_strength"),
                        )
                    )
        finally:
            wb.close()
        return records

    def extract_records_from_bytes(self, content: bytes, filename: str = "upload.xlsx") -> List[NfxInvestorRecord]:
        """Parse xlsx bytes (from API upload) into NfxInvestorRecord objects."""
        import io
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        records: List[NfxInvestorRecord] = []
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    continue

                header = all_rows[0]
                col_idx: dict[str, int] = {}
                for i, col in enumerate(header):
                    if col is None:
                        continue
                    key = str(col).strip().lower()
                    if key in self._COL_MAP:
                        col_idx[self._COL_MAP[key]] = i

                if "investor_name" not in col_idx:
                    continue

                for row in all_rows[1:]:
                    def get(field: str) -> Optional[str]:
                        idx = col_idx.get(field)
                        if idx is None or idx >= len(row):
                            return None
                        return _clean(row[idx])

                    name = get("investor_name")
                    if not name:
                        continue

                    records.append(
                        NfxInvestorRecord(
                            investor_name=name,
                            firm_name=get("firm_name"),
                            nfx_url=get("nfx_url"),
                            sweet_spot=_clean_dollar(get("sweet_spot")),
                            check_min=_clean_dollar(get("check_min")),
                            check_max=_clean_dollar(get("check_max")),
                            locations=get("locations"),
                            intro_source=get("intro_source"),
                            intro_strength=get("intro_strength"),
                        )
                    )
        finally:
            wb.close()
        return records

    def extract_raw(self, source: Source) -> Iterator[RawRecord]:
        """Yield standard RawRecord objects for entities_raw ingestion."""
        for record in self.extract_records(source):
            row_dict = record.model_dump()
            row_dict["_source_platform"] = "nfx_signal"
            content_hash = hash_content(row_dict)
            source_offset = f"name:{record.investor_name}"
            source_record_id = make_source_record_id(
                source.relative_path, source_offset, content_hash
            )
            yield RawRecord(
                source_record_id=source_record_id,
                source_file=source.relative_path,
                source_type=self.source_type,
                source_offset=source_offset,
                content_hash=content_hash,
                raw_content=row_dict,
                schema_version=self.schema_version,
            )

    def manifest(self, source: Source) -> SourceManifest:
        records = list(self.extract_raw(source))
        return SourceManifest(
            source_file=source.relative_path,
            source_type=self.source_type,
            file_hash=source.file_hash,
            record_count=len(records),
            quality_flags={"source_platform": "nfx_signal"},
        )
